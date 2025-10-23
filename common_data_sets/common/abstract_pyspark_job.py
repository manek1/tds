from abc import ABC, abstractmethod
from pyspark import SparkContext
import logging
from pyspark.sql.types import *
from .spark_common_functions import is_non_empty
import awswrangler as wr
import io
import boto3
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook


class AbstractPySparkJob(ABC):
    def __init__(self) -> None:
        self.sc: SparkContext = SparkContext.getOrCreate()
        self.spark = self.init_spark_job()
        self.columns = None

    @property
    @abstractmethod
    def app_name(self) -> str:
        pass

    @property
    @abstractmethod
    def job_type(self) -> str:
        pass

    @abstractmethod
    def _execute(self):
        pass

    @abstractmethod
    def init_spark_job(self):
        pass

    def read_s3_file(self, s3_path: str, schema: StructType = None, file_format: str = "parquet", options: dict = None):
        """
        :param options: read options
        :param file_format: Storage type in string format
        :param schema: Schema of data StructType
        :param s3_path: S3 path
        """

        reader = self.spark.read

        if options:
            for key, value in options.items():
                reader = reader.option(key, value)

        if schema:
            reader = reader.schema(schema)

        if file_format.lower() == 'parquet':
            return reader.parquet(s3_path)

        else:
            return reader.format(file_format).option("header", "true").load(s3_path)

    def write_df_to_s3(self, df, s3_path, out_format: str = 'parquet', mode_type: str = 'overwrite', num_of_files=-1,
                       partition=None, redshift_partition_col=None, **properties) -> None:
        """
        :param df: Spark DataFrame
        :param s3_path: S3 path
        :param out_format: output format
        :param overwrite: bool - True for overwrite else False
        :param num_of_files: Number output files
        :param partition: list of partition columns
        :return: None
        """
        if redshift_partition_col:
            df = df = df.withColumn(f"{redshift_partition_col}_val", df[redshift_partition_col])
        if is_non_empty(df):
            if num_of_files > 0:
                df = df.coalesce(num_of_files)
            write = df.write.format(out_format)
            if partition:
                write = write.partitionBy(partition)
            if mode_type == "overwrite":
                write = write.mode('overwrite')
            if mode_type == "append":
                write = write.mode('append')
            if properties:
                for name, val in properties.items():
                    write.option(name, val)

            write.save(s3_path)
            self.logger.info(f"Inserted the {df.count()} records in {s3_path} data")
        else:
            self.logger.info("Data Frame is empty, No data is written")

    def read_sql(self, sql_query):
        """
        :param sql_query: query required to run against source database
        """
        sql_reader = self.spark
        try:
            input_df = sql_reader.sql(sql_query)
            return input_df

        except:
            raise Exception(
                f"Reading data from athena tables  for query {sql_query} failed"
            )

    @property
    def logger(self):
        MSG_FORMAT = "%(asctime)s %(levelname)s %(name)s: %(message)s"
        DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
        logging.basicConfig(format=MSG_FORMAT, datefmt=DATETIME_FORMAT)
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)

        return logger

    def read_s3_files_using_awswrangler(self, s3_path, file_format):
        """
        :param s3_path: s3 path where source data present
        """
        try:
            if file_format == "parquet":
                files = wr.s3.list_objects(s3_path)
                print(files)
                input_dfs = [wr.s3.read_parquet(path=file) for file in files]
                input_df = pd.concat(input_dfs, ignore_index=True)
            elif file_format == "csv":
                input_df = wr.s3.read_csv(s3_path)
            elif file_format == "excel":
                input_df = wr.s3.read_excel(s3_path, keep_default_na=False)
            else:
                raise Exception(
                    f"provided file format {file_format} is not supported"
                )
        except:
            raise Exception(
                f"error in reading source df from path {s3_path}"
            )
        return input_df

    def pandas_to_spark(self, pandas_df):
        """
        Converts a Pandas DataFrame to a PySpark DataFrame with automatic schema inference.
        """

        return self.spark.createDataFrame(pandas_df)

    def write_s3_files_using_awswrangler(self, df, s3_path):
        """
        :param s3_path: s3 path where source data present
        """
        try:
            wr.s3.to_excel(df, path=s3_path, index=False)
        except:
            raise Exception(
                f"error in writing data to path  {s3_path}"
            )

    def write_xlsm_files(self, df, s3_input_path, s3_output_path):
        """
        :param s3_path: s3 path where source data present
        """
        s3_client = boto3.client("s3")
        s3_output_path = s3_output_path.replace('s3://', '')
        bucket_name = s3_output_path.split('/')[0]
        xlsm_output_key = '/'.join(s3_output_path.split('/')[1:])
        xlsm_input_path = s3_input_path.replace('s3://', '')
        xlsm_input_key = '/'.join(xlsm_input_path.split('/')[1:])

        self.logger.info(f'xlsx output path is {xlsm_output_key}')
        self.logger.info(f'xlsx output path is {xlsm_input_key}')
        self.logger.info(f'bucket is {bucket_name}')

        try:
            # Step 1: Download the `.xlsm` template from S3 (which contains macros)
            response = s3_client.get_object(Bucket=bucket_name, Key=xlsm_input_key)
            xlsm_content = response["Body"].read()

            # Step 2: Load the `.xlsm` file with `keep_vba=True` to retain macros
            input_stream = BytesIO(xlsm_content)
            wb = load_workbook(input_stream, keep_vba=True)

            # Step 3: Select the active worksheet (or a specific one, e.g., wb["Sheet1"])
            ws = wb.active

            # Step 4: Clear existing data in the worksheet (optional, if you want a clean sheet)
            ws.delete_cols(1, ws.max_column)  # Deletes all columns
            ws.delete_rows(1, ws.max_row)  # Deletes all rows

            # Step 5: Write new column headers (df_final column names)
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)  # Write column headers

            # Step 6: Write DataFrame (`df_final`) values into the worksheet
            for row_idx, row in enumerate(df.itertuples(index=False, name=None),
                                          start=2):  # Start from row 2 (skip headers)
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Step 7: Save the updated file back to a BytesIO buffer
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Step 8: Upload the updated `.xlsm` file back to S3
            s3_client.put_object(Bucket=bucket_name, Key=xlsm_output_key, Body=output.getvalue())

            self.logger.info(f"Saved {xlsm_output_key} with only df_final data while keeping macros.")
        except:
            raise Exception(
                f"error in writing data to path  {xlsm_output_key}"
            )
